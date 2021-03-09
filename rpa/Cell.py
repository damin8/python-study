from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Favian's Sheet"

ws["A1"] = 1
print(ws["A1"])
print(ws["A1"].value)  # 값이 없을 때는 "None" 출력

c = ws.cell(column=3, row=1, value=10)  # == ws["C1"].value = 10
print(c.value)
