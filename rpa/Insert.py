from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")
ws = wb.active

ws.iter_rows(8) # 8번째 줄에 1줄 추가
ws.iter_rows(8, 5) # 8번째 줄 위치에 5줄 추가

ws.insert_cols(2) # B번째 열이 비워짐 (새로운 빈 열이 추가)
ws.insert_cols(2, 3) # B번째 열 위치에 3줄 추가