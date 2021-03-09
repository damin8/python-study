from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet("Favian's Sheet")  # 새로운 Sheet 생성
ws.sheet_properties.tabColor = "ff66ff"

temp_ws = wb["Favian's Sheet"]

print(wb.sheetnames)

copied_sheet = wb.copy_worksheet(temp_ws)
copied_sheet.title = "Copied Sheet"

wb.save("sample.xlsx")
